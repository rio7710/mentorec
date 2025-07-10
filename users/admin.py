from django.contrib import admin, messages
from django.contrib.auth.admin import UserAdmin
from django.urls import path, reverse
from django.utils.html import format_html
from django.shortcuts import render, redirect
from django.db import transaction
from django.http import HttpResponse
import io
import openpyxl
from .models import User, Affiliation
from .forms import FileUploadForm
from instructors.models import InstructorProfile # 강사 프로필 생성을 위해 import

@admin.register(Affiliation)
class AffiliationAdmin(admin.ModelAdmin):
    list_display = ('name', 'type')
    list_filter = ('type',)
    search_fields = ('name',)

class CustomUserAdmin(UserAdmin):
    # 관리자 페이지 상세 정보 화면 필드 구성
    fieldsets = (
        (None, {"fields": ("username", "password")}),
        ("개인 정보", {"fields": ("first_name", "last_name", "email")}),
        ("추가 정보", {"fields": ("role", "instructor_profile_link", "affiliations")}),
        (
            "권한",
            {"fields": ("is_active", "is_staff", "is_superuser", "groups", "user_permissions")},
        ),
        ("주요 날짜", {"fields": ("last_login", "date_joined")}),
    )
    # 사용자 목록 화면에 보여질 필드
    list_display = ('username', 'email', 'is_staff', 'role')
    # 읽기 전용 필드
    readonly_fields = ('instructor_profile_link',)
    # 다대다 관계 필드를 위한 위젯
    filter_horizontal = ('affiliations', 'groups', 'user_permissions')
    # 사용자 목록 페이지에 커스텀 버튼을 추가하기 위한 템플릿
    change_list_template = "admin/users/user_change_list.html"

    def get_urls(self):
        """커스텀 URL 추가"""
        urls = super().get_urls()
        custom_urls = [
            path('bulk-upload/', self.admin_site.admin_view(self.bulk_upload_view), name='users_user_bulk_upload'),
            path('download-template/', self.admin_site.admin_view(self.download_excel_template_view), name='users_user_download_template'),
        ]
        return custom_urls + urls

    def download_excel_template_view(self, request):
        """엑셀 템플릿 다운로드 뷰"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "User Upload Template"
        headers = ["username", "email", "role"]
        sheet.append(headers)

        file_stream = io.BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)

        response = HttpResponse(
            file_stream,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="user_upload_template.xlsx"'
        return response

    def bulk_upload_view(self, request):
        """엑셀 파일로 사용자 일괄 등록하는 뷰"""
        if request.method == 'POST':
            form = FileUploadForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active
                
                errors = []
                new_users_data = []

                for row_num, row in enumerate(sheet.iter_rows(min_row=2), 2):
                    if len(row) < 3:
                        errors.append(f"{row_num}행: 데이터가 부족합니다. (아이디, 이메일, 역할 필요)")
                        continue
                        
                    username, email, role_str = [cell.value for cell in row[:3]]

                    if not all([username, email, role_str]):
                        errors.append(f"{row_num}행: 필수 항목(아이디, 이메일, 역할)이 비어있습니다.")
                        continue
                    if User.objects.filter(username=username).exists():
                        errors.append(f"{row_num}행: 아이디 '{username}'가 이미 존재합니다.")
                        continue
                    if User.objects.filter(email=email).exists():
                        errors.append(f"{row_num}행: 이메일 '{email}'이 이미 존재합니다.")
                        continue
                    
                    role_values = [choice[0] for choice in User.Role.choices]
                    if str(role_str).upper() not in role_values:
                        errors.append(f"{row_num}행: 역할(Role) 값 '{role_str}'이 잘못되었습니다. (STAFF, INSTRUCTOR, USER 중 하나여야 합니다)")
                        continue

                    new_users_data.append({'username': username, 'email': email, 'role': str(role_str).upper()})

                if errors:
                    for error in errors:
                        messages.error(request, error)
                else:
                    try:
                        with transaction.atomic():
                            for user_data in new_users_data:
                                user = User(username=user_data['username'], email=user_data['email'], role=user_data['role'])
                                user.set_password('password123')
                                user.save()
                                
                                if user.role == User.Role.INSTRUCTOR:
                                    InstructorProfile.objects.create(user=user)
                        messages.success(request, f"{len(new_users_data)}명의 사용자를 성공적으로 추가했습니다.")
                    except Exception as e:
                        messages.error(request, f"데이터베이스 저장 중 오류 발생: {e}")

                return redirect('.')
        else:
            form = FileUploadForm()
        
        context = dict(
           self.admin_site.each_context(request),
           form=form,
           title="엑셀로 사용자 일괄 등록"
        )
        return render(request, 'admin/users/bulk_upload.html', context)
    
    def instructor_profile_link(self, obj):
        """강사 프로필 바로가기 링크 생성 함수"""
        if obj.role == User.Role.INSTRUCTOR:
            try:
                profile_id = obj.instructorprofile.user_id
                url = reverse('admin:instructors_instructorprofile_change', args=[profile_id])
                return format_html('<a href="{}">✅ 강사 프로필 바로가기</a>', url)
            except User.instructorprofile.RelatedObjectDoesNotExist:
                url = reverse('admin:instructors_instructorprofile_add') + f'?user={obj.id}'
                return format_html('<a href="{}">➕ 강사 프로필 생성하기</a>', url)
        return "- (강사가 아님) -"
    
    instructor_profile_link.short_description = '프로필 링크'

# User 모델을 CustomUserAdmin으로 등록
admin.site.register(User, CustomUserAdmin)